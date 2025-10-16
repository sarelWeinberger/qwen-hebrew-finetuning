import fsspec
import gzip
import shutil
import os
import json
from collections import Counter, defaultdict
import tempfile
import pandas as pd
from datetime import datetime
import random
import boto3
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import re
from openpyxl.utils.exceptions import IllegalCharacterError

def download_and_analyze_files(file_configs, n_examples=5):
    """
    Download and analyze multiple S3 files for filter reasons and word counts.
    Also collect examples for each filter reason.
    
    Args:
        file_configs: List of dictionaries with 's3_path' and 'filter_group' keys
        n_examples: Number of examples to collect per filter reason
    """
    
    # Create S3 filesystem
    fs = fsspec.filesystem('s3')
    
    # Counters for aggregated results
    all_filter_reason_counts = Counter()
    all_filter_reason_word_sums = defaultdict(int)
    filter_group_stats = defaultdict(lambda: {
        'filter_counts': Counter(),
        'word_sums': defaultdict(int),
        'total_lines': 0
    })
    
    # Store examples for each filter reason
    filter_reason_examples = defaultdict(list)
    
    for config in file_configs:
        s3_path = config['s3_path']
        filter_group = config['filter_group']
        
        print(f"Processing {filter_group}: {s3_path}")
        
        # Create temporary files
        with tempfile.NamedTemporaryFile(suffix='.gz', delete=False) as temp_gz:
            temp_gz_path = temp_gz.name
        
        with tempfile.NamedTemporaryFile(suffix='.jsonl', delete=False) as temp_jsonl:
            temp_jsonl_path = temp_jsonl.name
        
        try:
            # Download the compressed file
            fs.get(s3_path, temp_gz_path)
            
            # Extract the .gz file
            with gzip.open(temp_gz_path, 'rb') as f_in:
                with open(temp_jsonl_path, 'wb') as f_out:
                    shutil.copyfileobj(f_in, f_out)
            
            # Analyze the JSONL file
            line_count = 0
            with open(temp_jsonl_path, 'r', encoding='utf-8') as file:
                for line_num, line in enumerate(file, 1):
                    try:
                        # Parse each JSON line
                        data = json.loads(line.strip())
                        line_count += 1
                        
                        # Extract filter_reason and n_words from metadata
                        if 'metadata' in data:
                            metadata = data['metadata']
                            filter_reason = metadata.get('filter_reason', 'no_filter_reason')
                            n_words = int(metadata.get('n_words', '0'))
                            
                            # Update global counters
                            all_filter_reason_counts[filter_reason] += 1
                            all_filter_reason_word_sums[filter_reason] += n_words
                            
                            # Update filter group counters
                            filter_group_stats[filter_group]['filter_counts'][filter_reason] += 1
                            filter_group_stats[filter_group]['word_sums'][filter_reason] += n_words
                            
                            # Collect examples (with reservoir sampling for memory efficiency)
                            if len(filter_reason_examples[filter_reason]) < n_examples:
                                # Add example if we haven't reached the limit
                                example = {
                                    'filter_reason': filter_reason,
                                    'text': data.get('text', ''),  # Keep full text, don't truncate here
                                    'text_preview': data.get('text', '')[:200] + '...' if len(data.get('text', '')) > 200 else data.get('text', ''),
                                    'n_words': n_words,
                                    'file_path': metadata.get('file_path', ''),
                                    'filter_group': filter_group,
                                    'record_id': data.get('id', ''),
                                    'full_metadata': json.dumps(metadata, indent=2)  # Pretty format JSON
                                }
                                filter_reason_examples[filter_reason].append(example)
                            else:
                                # Use reservoir sampling to potentially replace an existing example
                                current_count = all_filter_reason_counts[filter_reason]
                                if random.randint(1, current_count) <= n_examples:
                                    replace_idx = random.randint(0, n_examples - 1)
                                    example = {
                                        'filter_reason': filter_reason,
                                        'text': data.get('text', ''),
                                        'text_preview': data.get('text', '')[:200] + '...' if len(data.get('text', '')) > 200 else data.get('text', ''),
                                        'n_words': n_words,
                                        'file_path': metadata.get('file_path', ''),
                                        'filter_group': filter_group,
                                        'record_id': data.get('id', ''),
                                        'full_metadata': json.dumps(metadata, indent=2)
                                    }
                                    filter_reason_examples[filter_reason][replace_idx] = example
                        else:
                            # Handle lines without metadata
                            filter_reason = 'no_metadata'
                            all_filter_reason_counts[filter_reason] += 1
                            filter_group_stats[filter_group]['filter_counts'][filter_reason] += 1
                            
                            # Add example for no_metadata case
                            if len(filter_reason_examples[filter_reason]) < n_examples:
                                example = {
                                    'filter_reason': filter_reason,
                                    'text': data.get('text', ''),
                                    'text_preview': data.get('text', '')[:200] + '...' if len(data.get('text', '')) > 200 else data.get('text', ''),
                                    'n_words': 0,
                                    'file_path': '',
                                    'filter_group': filter_group,
                                    'record_id': data.get('id', ''),
                                    'full_metadata': 'No metadata available'
                                }
                                filter_reason_examples[filter_reason].append(example)
                            
                    except json.JSONDecodeError as e:
                        print(f"    Error parsing line {line_num}: {e}")
                        continue
                    except ValueError as e:
                        print(f"    Error converting n_words on line {line_num}: {e}")
                        continue
            
            filter_group_stats[filter_group]['total_lines'] = line_count
            print(f"  Processed {line_count:,} lines")
            
        finally:
            # Clean up temporary files
            if os.path.exists(temp_gz_path):
                os.remove(temp_gz_path)
            if os.path.exists(temp_jsonl_path):
                os.remove(temp_jsonl_path)
    
    return all_filter_reason_counts, all_filter_reason_word_sums, filter_group_stats, filter_reason_examples

def upload_file_to_s3(local_file_path, bucket_name, s3_key, content_type='text/plain'):
    """Helper function to upload a file to S3"""
    s3_client = boto3.client('s3')
    try:
        s3_client.upload_file(
            local_file_path,
            bucket_name,
            s3_key,
            ExtraArgs={
                'ContentType': content_type,
                'Metadata': {
                    'created_by': 'python_script',
                    'upload_time': datetime.now().isoformat()
                }
            }
        )
        return True
    except Exception as e:
        print(f"Error uploading {local_file_path} to S3: {e}\nBucket name: {bucket_name}, S3 key: {s3_key}")
        return False

def upload_directory_to_s3(local_dir, bucket_name, s3_prefix):
    """Upload all files in a directory to S3"""
    s3_client = boto3.client('s3')
    uploaded_files = []
    
    for root, dirs, files in os.walk(local_dir):
        for file in files:
            local_path = os.path.join(root, file)
            # Create relative path for S3
            relative_path = os.path.relpath(local_path, local_dir)
            s3_key = f"{s3_prefix}/{relative_path}".replace('\\', '/')  # Handle Windows paths
            
            if upload_file_to_s3(local_path, bucket_name, s3_key, 'text/plain'):
                uploaded_files.append(f"s3://{bucket_name}/{s3_key}")
    
    return uploaded_files

def clean_text_for_excel(text):
    """Clean text to remove characters that are illegal in Excel"""
    if not isinstance(text, str):
        return text
    
    # Remove control characters (0x00-0x1F except tab, newline, carriage return)
    # and other problematic characters
    cleaned = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', text)
    
    # Remove BOM and other Unicode formatting characters that can cause issues
    cleaned = re.sub(r'[\uFEFF\uFFFE\uFFFF]', '', cleaned)
    
    # Remove any remaining control characters that might cause issues
    cleaned = ''.join(char for char in cleaned if ord(char) >= 32 or char in '\t\n\r')
    
    # Limit length if too long (Excel's character limit per cell is 32,767)
    if len(cleaned) > 32767:
        cleaned = cleaned[:32767]
    
    return cleaned

def clean_dataframe_for_excel(df):
    """Clean all string columns in a dataframe for Excel export"""
    df_cleaned = df.copy()
    
    for col in df_cleaned.columns:
        if df_cleaned[col].dtype == 'object':  # String columns
            df_cleaned[col] = df_cleaned[col].apply(clean_text_for_excel)
    
    return df_cleaned

def safe_to_excel(df, writer, sheet_name, index=False, max_retries=2):
    """Safely write dataframe to Excel with error handling and cleaning"""
    for attempt in range(max_retries + 1):
        try:
            if attempt == 0:
                # First attempt: try with original data
                df.to_excel(writer, sheet_name=sheet_name, index=index)
                return True
            else:
                # Subsequent attempts: clean the data
                print(f"Attempt {attempt + 1}: Cleaning data for sheet '{sheet_name}'...")
                df_cleaned = clean_dataframe_for_excel(df)
                df_cleaned.to_excel(writer, sheet_name=sheet_name, index=index)
                print(f"Successfully wrote sheet '{sheet_name}' after cleaning")
                return True
                
        except IllegalCharacterError as e:
            print(f"IllegalCharacterError in sheet '{sheet_name}' (attempt {attempt + 1}): {e}")
            if attempt == max_retries:
                print(f"Failed to write sheet '{sheet_name}' after {max_retries + 1} attempts")
                return False
        except Exception as e:
            print(f"Unexpected error writing sheet '{sheet_name}': {e}")
            return False
    
    return False

def save_to_excel(filter_counts, word_sums, group_stats, examples, dataset_name, bucket_name, root_path, output_file="filter_analysis_results.xlsx"):
    """Save all results to Excel with multiple sheets and proper formatting"""
    filename = f'{dataset_name}_{output_file}'
    
    try:
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            total_word_sum = sum(list([a[1] for a in word_sums.items()]))
            
            # Sheet 1: Overall Summary
            summary_data = []
            for reason, count in filter_counts.most_common():
                total_words = word_sums[reason]
                avg_words = total_words / count if count > 0 else 0
                summary_data.append({
                    'Filter Reason': reason,
                    'Count': count,
                    'Total Words': total_words,
                    'Average Words': round(avg_words, 2),
                    'Count Percentage': round((count / sum(filter_counts.values())) * 100, 2),
                    'Words Percentage': round((total_words / total_word_sum) * 100, 2),
                })
            
            summary_df = pd.DataFrame(summary_data)
            safe_to_excel(summary_df, writer, 'Overall Summary')
            
            # Sheet 2: Group Statistics
            group_data = []
            for group_name, stats in group_stats.items():
                for reason, count in stats['filter_counts'].items():
                    total_words = stats['word_sums'][reason]
                    avg_words = total_words / count if count > 0 else 0
                    group_data.append({
                        'Filter Group': group_name,
                        'Filter Reason': reason,
                        'Count': count,
                        'Total Words': total_words,
                        'Average Words': round(avg_words, 2),
                        'Group Total Lines': stats['total_lines']
                    })
            
            group_df = pd.DataFrame(group_data)
            safe_to_excel(group_df, writer, 'Group Statistics')
            
            # Sheet 3: Examples Preview (with short text)
            examples_preview_data = []
            for reason, example_list in examples.items():
                for i, example in enumerate(example_list, 1):
                    examples_preview_data.append({
                        'Filter Reason': reason,
                        'Example #': i,
                        'Text Preview': example['text_preview'],
                        'Word Count': example['n_words'],
                        'Record ID': example['record_id'],
                        'Filter Group': example['filter_group'],
                        'File Path': example['file_path']
                    })
            
            examples_preview_df = pd.DataFrame(examples_preview_data)
            safe_to_excel(examples_preview_df, writer, 'Examples Preview')
            
            # Sheet 4: Full Examples (with complete text and metadata)
            examples_full_data = []
            for reason, example_list in examples.items():
                for i, example in enumerate(example_list, 1):
                    examples_full_data.append({
                        'Filter Reason': reason,
                        'Example #': i,
                        'Full Text': example['text'],
                        'Word Count': example['n_words'],
                        'Record ID': example['record_id'],
                        'Filter Group': example['filter_group'],
                        'File Path': example['file_path'],
                        'Full Metadata': example['full_metadata']
                    })
            
            examples_full_df = pd.DataFrame(examples_full_data)
            safe_to_excel(examples_full_df, writer, 'Full Examples')
            
            # Sheet 5: Filter Reason Counts
            reason_counts_df = pd.DataFrame([
                {'Filter Reason': reason, 'Count': count} 
                for reason, count in filter_counts.most_common()
            ])
            safe_to_excel(reason_counts_df, writer, 'Filter Counts')
            
            # Format the worksheets for better readability
            workbook = writer.book
            
            # Format Examples sheets
            for sheet_name in ['Examples Preview', 'Full Examples']:
                if sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                    
                    # Set column widths
                    column_widths = {
                        'A': 20,  # Filter Reason
                        'B': 12,  # Example #
                        'C': 80,  # Text/Text Preview
                        'D': 12,  # Word Count
                        'E': 20,  # Record ID
                        'F': 15,  # Filter Group
                        'G': 30,  # File Path
                        'H': 50   # Full Metadata (if exists)
                    }
                    
                    for col, width in column_widths.items():
                        worksheet.column_dimensions[col].width = width
                    
                    # Set text wrap for text columns
                    for row in worksheet.iter_rows(min_row=2):  # Skip header
                        for cell in row:
                            if cell.column_letter in ['C', 'G', 'H']:  # Text columns
                                cell.alignment = Alignment(wrap_text=True, vertical='top')
                            else:
                                cell.alignment = Alignment(vertical='top')
                    
                    # Set row height for better readability
                    for row_num in range(2, worksheet.max_row + 1):
                        worksheet.row_dimensions[row_num].height = 60
        
        # Upload Excel file to S3
        s3_key = f'{root_path}/{dataset_name}/logs/{filename}'
        if upload_file_to_s3(filename, bucket_name, s3_key, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'):
            print(f"\nExcel file uploaded to S3: s3://{bucket_name}/{s3_key}")
        else:
            print(f"Failed to upload Excel file to S3")
    
    except Exception as e:
        print(f"Error creating Excel file: {e}")
        # Fallback: try saving as CSV
        try:
            print("Attempting to save as CSV instead...")
            csv_filename = filename.replace('.xlsx', '.csv')
            
            # Combine all data into one CSV
            all_data = []
            for reason, example_list in examples.items():
                for i, example in enumerate(example_list, 1):
                    all_data.append({
                        'Filter Reason': clean_text_for_excel(reason),
                        'Example #': i,
                        'Full Text': clean_text_for_excel(example['text']),
                        'Text Preview': clean_text_for_excel(example['text_preview']),
                        'Word Count': example['n_words'],
                        'Record ID': clean_text_for_excel(str(example['record_id'])),
                        'Filter Group': clean_text_for_excel(example['filter_group']),
                        'File Path': clean_text_for_excel(example['file_path']),
                        'Full Metadata': clean_text_for_excel(str(example['full_metadata']))
                    })
            
            csv_df = pd.DataFrame(all_data)
            csv_df.to_csv(csv_filename, index=False, encoding='utf-8-sig')
            print(f"CSV fallback saved to: {csv_filename}")
            
            # Try to upload CSV to S3
            s3_key_csv = f'{root_path}/{dataset_name}/logs/{csv_filename}'
            if upload_file_to_s3(csv_filename, bucket_name, s3_key_csv, 'text/csv'):
                print(f"CSV file uploaded to S3: s3://{bucket_name}/{s3_key_csv}")
            
            return csv_filename
            
        except Exception as csv_error:
            print(f"Failed to save CSV fallback: {csv_error}")
            raise e
    
    print(f"\nExcel results saved locally to: {filename}")
    return filename

def save_examples_to_files_and_s3(examples, dataset_name, bucket_name,root_path):
    """Save examples to local files and upload to S3"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S') 
    # 1. Save as JSON for programmatic access
    json_filename = f'{dataset_name}_examples_{timestamp}.json'
    with open(json_filename, 'w', encoding='utf-8') as f:
        json.dump(examples, f, indent=2, ensure_ascii=False)
    
    # Upload JSON to S3
    json_s3_key = f'{root_path}/{dataset_name}/logs/examples/{json_filename}'
    json_uploaded = upload_file_to_s3(json_filename, bucket_name, json_s3_key, 'application/json')
    
    # 2. Save as text files organized by filter reason
    examples_dir = f'filtering_examples/content_filtering_{timestamp}/{dataset_name}'
    os.makedirs(examples_dir, exist_ok=True)
    
    # Create a summary file
    summary_file = os.path.join(examples_dir, '_SUMMARY.txt')
    with open(summary_file, 'w', encoding='utf-8') as f:
        f.write(f"Examples Analysis Summary for {dataset_name}\n")
        f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=" * 80 + "\n\n")
        f.write(f"Total filter reasons: {len(examples)}\n")
        f.write(f"Total examples: {sum(len(examples_list) for examples_list in examples.values())}\n\n")
        
        f.write("Filter reasons and example counts:\n")
        f.write("-" * 40 + "\n")
        for filter_reason, example_list in sorted(examples.items()):
            f.write(f"{filter_reason}: {len(example_list)} examples\n")
    
    # Create individual files for each filter reason
    txt_files_created = []
    for filter_reason, example_list in examples.items():
        # Create safe filename
        safe_filter_name = "".join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in filter_reason)
        txt_filename = os.path.join(examples_dir, f'{safe_filter_name}.txt')
        txt_files_created.append(txt_filename)
        
        with open(txt_filename, 'w', encoding='utf-8') as f:
            f.write(f"Filter Reason: {filter_reason}\n")
            f.write(f"Number of examples: {len(example_list)}\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("=" * 80 + "\n\n")
            
            for i, example in enumerate(example_list, 1):
                f.write(f"EXAMPLE {i}\n")
                f.write("-" * 40 + "\n")
                f.write(f"Record ID: {example['record_id']}\n")
                f.write(f"Filter Group: {example['filter_group']}\n")
                f.write(f"Word Count: {example['n_words']}\n")
                f.write(f"File Path: {example['file_path']}\n")
                f.write(f"\nText:\n{example['text']}\n")
                f.write(f"\nMetadata:\n{example['full_metadata']}\n")
                f.write("\n" + "=" * 80 + "\n\n")
    
    # 3. Upload all text files to S3
    s3_prefix = f'{root_path}/{dataset_name}/logs/{examples_dir}'
    uploaded_files = upload_directory_to_s3(examples_dir, bucket_name, s3_prefix)
    
    # 4. Create and upload a ZIP file for easy download
    import zipfile
    zip_filename = f'{examples_dir}.zip'
    with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
        # Add all files in the examples directory
        for root, dirs, files in os.walk(examples_dir):
            for file in files:
                file_path = os.path.join(root, file)
                arcname = os.path.relpath(file_path, examples_dir)
                zipf.write(file_path, arcname)
        
        # Also add the JSON file to the zip
        zipf.write(json_filename, f'{json_filename}')
    
    # Upload ZIP to S3
    zip_s3_key = f'{root_path}/{dataset_name}/logs/{zip_filename}'
    zip_uploaded = upload_file_to_s3(zip_filename, bucket_name, zip_s3_key, 'application/zip')
    
    print(f"\nExamples saved locally:")
    print(f"  - JSON: {json_filename}")
    print(f"  - Text files directory: {examples_dir}/")
    print(f"  - ZIP file: {zip_filename}")
    
    print(f"\nExamples uploaded to S3:")
    if json_uploaded:
        print(f"  - JSON: s3://{bucket_name}/{json_s3_key}")
    if uploaded_files:
        print(f"  - Text files: {len(uploaded_files)} files uploaded to s3://{bucket_name}/{s3_prefix}/")
        print(f"    Examples: {uploaded_files[:3]}{'...' if len(uploaded_files) > 3 else ''}")
    if zip_uploaded:
        print(f"  - ZIP file: s3://{bucket_name}/{zip_s3_key}")
    
    return {
        'local': {
            'json': json_filename,
            'directory': examples_dir,
            'zip': zip_filename
        },
        's3': {
            'json': f's3://{bucket_name}/{json_s3_key}' if json_uploaded else None,
            'directory': f's3://{bucket_name}/{s3_prefix}/' if uploaded_files else None,
            'zip': f's3://{bucket_name}/{zip_s3_key}' if zip_uploaded else None,
            'uploaded_files': uploaded_files
        }
    }

def print_results(filter_counts, word_sums, group_stats):
    """Print formatted results to console"""
    
    print("\n" + "="*80)
    print("OVERALL RESULTS")
    print("="*80)
    print(f"{'Filter Reason':<40} {'Count':<15} {'Total Words':<15} {'Avg Words':<10}")
    print("-" * 80)
    
    for reason in filter_counts.most_common():
        reason_name = reason[0]
        count = reason[1]
        total_words = word_sums[reason_name]
        avg_words = total_words / count if count > 0 else 0
        print(f"{reason_name:<40} {count:<15,} {total_words:<15,} {avg_words:<10.1f}")
    
    print(f"\nTotal lines across all files: {sum(filter_counts.values()):,}")
    print(f"Total words across all files: {sum(word_sums.values()):,}")
    
    # Print results by filter group
    print("\n" + "="*80)
    print("RESULTS BY FILTER GROUP")
    print("="*80)
    
    for group_name, stats in group_stats.items():
        print(f"\nFilter Group: {group_name}")
        print("-" * 50)
        print(f"Total lines: {stats['total_lines']:,}")
        print(f"{'Filter Reason':<30} {'Count':<10} {'Words':<12} {'Avg':<8}")
        print("-" * 50)
        
        for reason, count in stats['filter_counts'].most_common():
            total_words = stats['word_sums'][reason]
            avg_words = total_words / count if count > 0 else 0
            print(f"{reason:<30} {count:<10,} {total_words:<12,} {avg_words:<8.1f}")

def get_file_keys_only(bucket_name, prefix):
    """
    Simple version that returns only the file keys as a list
    """
    s3_client = boto3.client('s3')
    paginator = s3_client.get_paginator('list_objects_v2')
    
    file_keys = []
    
    try:
        for page in paginator.paginate(Bucket=bucket_name, Prefix=prefix):
            if 'Contents' in page:
                for obj in page['Contents']:
                    if not obj['Key'].endswith('/'):  # Skip directories
                        file_keys.append(obj['Key'])
        
        return file_keys
    
    except Exception as e:
        print(f"Error: {e}")
        return []
# Example usage
def filter_analysis_pipeline( dataset_name, bucket_name, root_path):
    # bucket_name = "gepeta-datasets"
    # dataset_name = "sefaria"  # "hebrew_tweets"
    
    # Define your files to process
    files_to_process = []
    for filt_name in ["min_doc", "gopher_qual", "repetition", "fineweb", "NO_FILTER"]:
        prefix = f"{root_path}/{dataset_name}/filtering/removed/{filt_name}/" if filt_name != "NO_FILTER" else f"{root_path}/{dataset_name}/filtering/output/"
        file_keys = get_file_keys_only(bucket_name, prefix)
        print(f"Found {len(file_keys)} files for filter '{filt_name}'")
        for key in file_keys:
            files_to_process.append({"s3_path": f"{bucket_name}/{key}", "filter_group": filt_name})

    # Configuration
    N_EXAMPLES = 30  # Number of examples to collect per filter reason
    OUTPUT_EXCEL_FILE = f"filter_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Process all files
    print("Starting analysis...")
    filter_counts, word_sums, group_stats, examples = download_and_analyze_files(
        files_to_process, 
        n_examples=N_EXAMPLES
    )
    
    # Print results to console
    print_results(filter_counts, word_sums, group_stats)
    
    # Save to Excel with improved formatting
    print(f"\nSaving {N_EXAMPLES} examples per filter reason to Excel...")
    save_to_excel(filter_counts, word_sums, group_stats, examples, dataset_name, bucket_name, root_path, OUTPUT_EXCEL_FILE)
    
    # Save examples to separate files for easier viewing
    print("\nSaving examples to separate files...")
    outputs_paths = save_examples_to_files_and_s3(examples, dataset_name, bucket_name, root_path)
    print(f"\n{dataset_name} Analysis complete!")
    print(f"Examples collected: {sum(len(examples_list) for examples_list in examples.values())}")
    print(f"Unique filter reasons found: {len(filter_counts)}")
    print(f"\nOutput files created:")
    print(f"  - Excel: {dataset_name}_{OUTPUT_EXCEL_FILE}")
    print(f"  - JSON: {outputs_paths['local']['json']}")
    print(f"  - Text files: {outputs_paths['local']['directory']}/")
    print(f"  - S3 Json file: {outputs_paths['s3']['json']}/")

if __name__ == "__main__":
    DATASETS = [ 'AllOfHEOscarData', 'AllHebNLI', 'AllTzenzuraData', 'BIU', 'BooksNLI2', 'COGNI', 'FineWeb2', 'GeektimeCorpus', 'HeC4-HF', 'HeC4DictaCombined', 'RAMA', 'SecuritiesAuthority', 'StateComptrollerReports', 'SupremeCourtOfIsrael', 'TauDigital', 'TauOCR', 'YifatDataBatch2', 'YifatDataBatch2-Round4', 'YifatDataBatch2-Round5', 'YifatDataRound2', 'YifatToCombine', 'YisraelHayomData', 'hebrew_tweets', 'kohelet', 'sefaria', 'wikipedia']
    for ds in DATASETS:
        filter_analysis_pipeline( dataset_name=ds, bucket_name="gepeta-datasets", root_path="processed_cleaned_filtered/run_5")